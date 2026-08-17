import sys
import tempfile
import types
import unittest
from pathlib import Path
from unittest.mock import patch


def _ensure_google_genai_stub():
    try:
        from google import genai as _genai  # noqa: F401
        from google.genai import types as _types  # noqa: F401
        return
    except Exception:
        google_module = sys.modules.get("google")
        if google_module is None:
            google_module = types.ModuleType("google")
            google_module.__path__ = []
            sys.modules["google"] = google_module

        genai_module = types.ModuleType("google.genai")
        genai_types_module = types.ModuleType("google.genai.types")
        genai_module.types = genai_types_module
        google_module.genai = genai_module

        sys.modules["google.genai"] = genai_module
        sys.modules["google.genai.types"] = genai_types_module


def _ensure_webview_stub():
    try:
        import webview  # noqa: F401
        return
    except Exception:
        webview_module = types.ModuleType("webview")
        webview_module.windows = []
        webview_module.create_window = lambda *args, **kwargs: None
        webview_module.start = lambda *args, **kwargs: None
        sys.modules["webview"] = webview_module


def _ensure_dotenv_stub():
    try:
        from dotenv import load_dotenv as _load_dotenv  # noqa: F401
        return
    except Exception:
        dotenv_module = types.ModuleType("dotenv")
        dotenv_module.load_dotenv = lambda *args, **kwargs: False
        sys.modules["dotenv"] = dotenv_module


def _ensure_requests_stub():
    try:
        import requests  # noqa: F401
        return
    except Exception:
        requests_module = types.ModuleType("requests")
        requests_module.get = lambda *args, **kwargs: None
        sys.modules["requests"] = requests_module


def _ensure_pydantic_stub():
    try:
        from pydantic import BaseModel as _BaseModel, Field as _Field  # noqa: F401
        return
    except Exception:
        pydantic_module = types.ModuleType("pydantic")

        class BaseModel:
            pass

        def Field(*args, **kwargs):
            if args:
                return args[0]
            return kwargs.get("default")

        pydantic_module.BaseModel = BaseModel
        pydantic_module.Field = Field
        sys.modules["pydantic"] = pydantic_module


_ensure_google_genai_stub()
_ensure_webview_stub()
_ensure_dotenv_stub()
_ensure_requests_stub()
_ensure_pydantic_stub()

import main as main_module
from main import Api
from openpyxl import load_workbook


class DeliverableExcelExportTests(unittest.TestCase):
    def setUp(self):
        self.api = Api.__new__(Api)

    def test_export_deliverables_excel_orders_rows_by_due_date_descending(self):
        with tempfile.TemporaryDirectory(prefix="deliverable-export-") as temp_dir:
            output_path = Path(temp_dir) / "deliverables.xlsx"
            payload = {
                "filePath": str(output_path),
                "entries": [
                    {
                        "projectId": "1003",
                        "projectName": "No Date",
                        "deliverableName": "Undated",
                        "due": "",
                        "statusText": "Waiting",
                    },
                    {
                        "projectId": "1002",
                        "projectName": "Past Project",
                        "deliverableName": "Past Deliverable",
                        "due": "01/15/2026",
                        "statusText": "Complete",
                    },
                    {
                        "projectId": "1001",
                        "projectName": "Future Project",
                        "deliverableName": "Future Deliverable",
                        "due": "12/31/2026",
                        "statusText": "Working",
                        "projectPath": r"M:\Gensler\BAC\2026\1001 Future Project",
                    },
                    {
                        "projectId": "1004",
                        "projectName": "Middle Project",
                        "deliverableName": "Middle Deliverable",
                        "due": "2026-06-01",
                        "hardDue": "06/15/2026",
                        "statusText": "Pending Review",
                    },
                ],
            }

            with patch.object(main_module.os, "startfile", create=True):
                result = self.api.export_deliverables_excel(payload)

            self.assertEqual("success", result["status"])
            self.assertEqual(str(output_path), result["path"])
            self.assertTrue(output_path.exists())

            workbook = load_workbook(output_path)
            self.addCleanup(workbook.close)
            worksheet = workbook["Deliverables"]

        self.assertEqual(
            (
                "Project ID",
                "Project Name",
                "Deliverable",
                "Internal Due",
                "Hard Deadline",
                "Status",
                "Project Path",
            ),
            tuple(cell.value for cell in worksheet[1]),
        )
        self.assertEqual("1001", worksheet["A2"].value)
        self.assertEqual("Future Project", worksheet["B2"].value)
        self.assertEqual("Future Deliverable", worksheet["C2"].value)
        self.assertEqual("Working", worksheet["F2"].value)
        self.assertEqual("1004", worksheet["A3"].value)
        self.assertEqual("1002", worksheet["A4"].value)
        self.assertEqual("1003", worksheet["A5"].value)
        self.assertEqual("mm/dd/yyyy", worksheet["D2"].number_format)
        self.assertEqual("01/15/2026", worksheet["D4"].value.strftime("%m/%d/%Y"))
        self.assertIsNone(worksheet["D5"].value)
        # Hard deadline lands in its own column, formatted as a real date.
        self.assertEqual("06/15/2026", worksheet["E3"].value.strftime("%m/%d/%Y"))
        self.assertEqual("mm/dd/yyyy", worksheet["E3"].number_format)
        self.assertIsNone(worksheet["E2"].value)
        # The project folder rides along in its own trailing column.
        self.assertEqual(
            r"M:\Gensler\BAC\2026\1001 Future Project", worksheet["G2"].value
        )
        # A project with no path recorded leaves the cell blank rather than erroring.
        self.assertIsNone(worksheet["G3"].value)

    def test_export_deliverables_excel_sorts_hard_only_deliverable_by_its_hard_date(self):
        with tempfile.TemporaryDirectory(prefix="deliverable-export-hard-") as temp_dir:
            output_path = Path(temp_dir) / "deliverables.xlsx"
            payload = {
                "filePath": str(output_path),
                "entries": [
                    {
                        "projectId": "2001",
                        "projectName": "Internal Only",
                        "deliverableName": "Internal Deliverable",
                        "due": "03/01/2026",
                        "statusText": "Waiting",
                    },
                    {
                        "projectId": "2002",
                        "projectName": "Hard Only",
                        "deliverableName": "Hard Deliverable",
                        "due": "",
                        "hardDue": "09/01/2026",
                        "statusText": "Waiting",
                    },
                    {
                        "projectId": "2003",
                        "projectName": "Undated",
                        "deliverableName": "Undated Deliverable",
                        "due": "",
                        "statusText": "Waiting",
                    },
                ],
            }

            with patch.object(main_module.os, "startfile", create=True):
                result = self.api.export_deliverables_excel(payload)

            self.assertEqual("success", result["status"])

            workbook = load_workbook(output_path)
            self.addCleanup(workbook.close)
            worksheet = workbook["Deliverables"]

        # A hard-only deliverable sorts on its hard date rather than falling to
        # the bottom with the genuinely undated rows.
        self.assertEqual("2002", worksheet["A2"].value)
        self.assertEqual("2001", worksheet["A3"].value)
        self.assertEqual("2003", worksheet["A4"].value)

    def test_export_deliverables_excel_preserves_unc_project_paths(self):
        unc = "\\\\acies.lan\\cachedrive\\projects\\Nelson\\BAC\\2024\\241039 BofA"
        with tempfile.TemporaryDirectory(prefix="deliverable-export-unc-") as temp_dir:
            output_path = Path(temp_dir) / "deliverables.xlsx"
            payload = {
                "filePath": str(output_path),
                "entries": [
                    {
                        "projectId": "241039",
                        "projectName": "BofA",
                        "deliverableName": "PCC",
                        "due": "05/08/2026",
                        "statusText": "On hold",
                        "projectPath": unc,
                    },
                ],
            }

            with patch.object(main_module.os, "startfile", create=True):
                result = self.api.export_deliverables_excel(payload)

            self.assertEqual("success", result["status"])
            workbook = load_workbook(output_path)
            self.addCleanup(workbook.close)
            worksheet = workbook["Deliverables"]

        # Both leading backslashes must survive or the path will not open.
        self.assertEqual(unc, worksheet["G2"].value)
        self.assertTrue(worksheet["G2"].value.startswith("\\\\"))

    def _export_with_summary(self, summary, prefix):
        """Exports one fixed entry, optionally with a summary, and returns the workbook."""
        with tempfile.TemporaryDirectory(prefix=prefix) as temp_dir:
            output_path = Path(temp_dir) / "deliverables.xlsx"
            payload = {
                "filePath": str(output_path),
                "entries": [
                    {
                        "projectId": "3001",
                        "projectName": "Summary Project",
                        "deliverableName": "Summary Deliverable",
                        "due": "04/01/2026",
                        "statusText": "None",
                    },
                ],
            }
            if summary is not None:
                payload["summary"] = summary

            with patch.object(main_module.os, "startfile", create=True):
                result = self.api.export_deliverables_excel(payload)

            self.assertEqual("success", result["status"])
            workbook = load_workbook(output_path)
            self.addCleanup(workbook.close)
            return workbook

    def test_export_deliverables_excel_omits_summary_sheet_when_summary_absent(self):
        # Regression lock: the summary extension must stay inert for the old payload.
        workbook = self._export_with_summary(None, "deliverable-export-nosummary-")
        self.assertEqual(["Deliverables"], workbook.sheetnames)

    def test_export_deliverables_excel_writes_status_summary_sheet(self):
        workbook = self._export_with_summary(
            {
                "headline": "Two hard deadlines are already blown.",
                "paragraphs": ["First para.", "Second para."],
                "generatedAt": "08/05/2026 09:00 AM",
                "scope": "incomplete",
                "deliverableCount": 2,
            },
            "deliverable-export-summary-",
        )

        self.assertIn("Status Summary", workbook.sheetnames)
        # The deliverable rows must remain the first and active sheet.
        self.assertEqual("Deliverables", workbook.sheetnames[0])
        self.assertEqual("3001", workbook["Deliverables"]["A2"].value)

        sheet = workbook["Status Summary"]
        self.assertEqual("AI Status Briefing", sheet["A1"].value)
        self.assertIn("08/05/2026", sheet["A2"].value)
        self.assertIn("2 deliverables", sheet["A2"].value)
        self.assertEqual("Two hard deadlines are already blown.", sheet["A4"].value)
        self.assertEqual("First para.", sheet["A6"].value)
        self.assertEqual("Second para.", sheet["A7"].value)
        self.assertTrue(sheet["A6"].alignment.wrap_text)

    def test_export_deliverables_excel_ignores_blank_summary(self):
        workbook = self._export_with_summary(
            {"headline": "", "paragraphs": ["", "   "]},
            "deliverable-export-blanksummary-",
        )
        self.assertEqual(["Deliverables"], workbook.sheetnames)


if __name__ == "__main__":
    unittest.main()
