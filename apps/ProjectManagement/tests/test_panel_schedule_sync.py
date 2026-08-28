"""
Unit tests for apps/ProjectManagement/panel_schedule_sync.py
"""

import os
import tempfile
import unittest
from copy import deepcopy
import openpyxl

from apps.ProjectManagement.database import (
    get_panel_schedule,
    get_panel_workbook_binding,
    init_db,
    upsert_project,
)
from apps.ProjectManagement.panel_schedule_sync import (
    PanelScheduleSyncService,
    calculate_phase_balance,
    parse_panel_sheet,
    parse_panel_workbook,
    sync_panel_workbook_to_db,
    COL_L_DESC,
    COL_L_TRIP,
    COL_L_POLE,
    COL_L_KVA,
    COL_R_DESC,
    COL_R_TRIP,
    COL_R_POLE,
    COL_R_KVA,
    COL_VOLTAGE
)


class TestPanelScheduleSync(unittest.TestCase):
    def setUp(self):
        self.fd, self.temp_db = tempfile.mkstemp(suffix=".db")
        os.close(self.fd)
        init_db(self.temp_db)

    def tearDown(self):
        if os.path.exists(self.temp_db):
            try:
                os.remove(self.temp_db)
            except OSError:
                pass

    def test_calculate_phase_balance(self):
        # Balanced 3-phase loads: A=2000, B=2000, C=2000
        circuits = [
            {"circuitNumber": 1, "phasePole": "A", "connectedVA": 2000.0},
            {"circuitNumber": 3, "phasePole": "B", "connectedVA": 2000.0},
            {"circuitNumber": 5, "phasePole": "C", "connectedVA": 2000.0}
        ]
        summary = calculate_phase_balance(circuits, phase_count=3)
        self.assertEqual(summary["phaseAConnectedVA"], 2000.0)
        self.assertEqual(summary["phaseBConnectedVA"], 2000.0)
        self.assertEqual(summary["phaseCConnectedVA"], 2000.0)
        self.assertEqual(summary["totalConnectedVA"], 6000.0)
        self.assertEqual(summary["unbalancePercentage"], 0.0)

        # Unbalanced 3-phase loads: A=3000, B=2000, C=1000. Avg = 2000. (3000-1000)/2000 * 100 = 100%
        circuits_unbalanced = [
            {"circuitNumber": 1, "phasePole": "A", "connectedVA": 3000.0},
            {"circuitNumber": 3, "phasePole": "B", "connectedVA": 2000.0},
            {"circuitNumber": 5, "phasePole": "C", "connectedVA": 1000.0}
        ]
        summary_unbal = calculate_phase_balance(circuits_unbalanced, phase_count=3)
        self.assertEqual(summary_unbal["unbalancePercentage"], 100.0)

    def test_parse_and_sync_panel_workbook(self):
        # Create a sample workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LP-1"

        ws["A3"] = "(E) PANEL 'LP-1'"
        ws[f"{COL_VOLTAGE}2"] = "120/208V"
        ws["G3"] = "225A"
        ws["K2"] = "4W"
        ws["K3"] = "3PH"
        ws["K4"] = "NEMA 1"
        ws["N2"] = "SURFACE"
        ws["N3"] = "10KAIC"

        # Row 8: Ckt 1 (L) and Ckt 2 (R)
        ws[f"{COL_L_DESC}8"] = "Lighting Corridor"
        ws[f"{COL_L_TRIP}8"] = "20"
        ws[f"{COL_L_POLE}8"] = "1"
        ws[f"{COL_L_KVA}8"] = "1.6"

        ws[f"{COL_R_DESC}8"] = "Receptacles Office"
        ws[f"{COL_R_TRIP}8"] = "20"
        ws[f"{COL_R_POLE}8"] = "1"
        ws[f"{COL_R_KVA}8"] = "1.8"

        temp_xlsx = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_xlsx_path = temp_xlsx.name
        temp_xlsx.close()

        try:
            wb.save(temp_xlsx_path)

            proj = upsert_project("24-999", "Sync Test Project", "C:/Projects/24-999", db_path=self.temp_db)
            synced = sync_panel_workbook_to_db(proj["id"], temp_xlsx_path, db_path=self.temp_db)

            self.assertEqual(len(synced), 1)
            panel = synced[0]
            self.assertEqual(panel["panel_name"], "LP-1")
            self.assertEqual(panel["voltage"], "120/208V")
            self.assertEqual(panel["main_bus_amps"], 225)
            self.assertGreater(len(panel["circuits"]), 0)

            # Check circuit 1
            ckt1 = next(c for c in panel["circuits"] if c["circuit_number"] == 1)
            self.assertEqual(ckt1["load_description"], "Lighting Corridor")
            self.assertEqual(ckt1["connected_va"], 1600.0)

            # Check circuit 2
            ckt2 = next(c for c in panel["circuits"] if c["circuit_number"] == 2)
            self.assertEqual(ckt2["load_description"], "Receptacles Office")
            self.assertEqual(ckt2["connected_va"], 1800.0)

        finally:
            if os.path.exists(temp_xlsx_path):
                os.remove(temp_xlsx_path)

    @staticmethod
    def _create_acies_plan_workbook(path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'PNL "LP-1"'
        ws["A2"] = '(E) PNL "LP-1"'
        ws["A3"] = "SECTION 1"
        ws["G1"] = "120 /"
        ws["H1"] = 208
        ws["I1"] = "V"
        ws["G2"] = 225
        ws["H2"] = "A"
        ws["G3"] = "225A"
        ws["H3"] = "MCB"
        ws["M1"] = 4
        ws["M2"] = 3
        ws["M3"] = "(E)"
        ws["P1"] = "SURFACE MOUNTED"
        ws["P2"] = "10KAIC"
        ws["P3"] = "EXISTING"
        ws["R2"] = "Feed from"
        ws["R3"] = '(E) PNL "MSB"'
        ws["A5"] = "CKT#"
        ws["A7"] = 1
        ws["B7"] = "(E)"
        ws["C7"] = "G"
        ws["D7"] = 1
        ws["E7"] = 20
        ws["F7"] = "RECEPTACLES"
        ws["K7"] = "=1+0.5"
        ws["L7"] = "A"
        ws["M7"] = 0.48
        ws["N7"] = "LIGHTING"
        ws["Q7"] = 1
        ws["R7"] = 20
        ws["S7"] = "C"
        ws["T7"] = "(E)"
        ws["U7"] = 2
        wb.save(path)
        wb.close()

    def test_attached_acies_layout_preserves_formula_inputs(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = os.path.join(temp_dir, "Panels.xlsx")
            self._create_acies_plan_workbook(workbook_path)

            service = PanelScheduleSyncService(
                db_path=self.temp_db,
                prefer_com=False,
            )
            state = service.begin_session("26-0036", workbook_path)

            self.assertEqual(state["status"], "success")
            self.assertEqual(len(state["panels"]), 1)
            panel = state["panels"][0]
            self.assertEqual(panel["panelName"], "LP-1")
            self.assertEqual(panel["displayName"], "LP-1 - SECTION 1")
            self.assertEqual(panel["voltage"], "120/208V")
            self.assertEqual(panel["circuits"][0]["connectedKvaInput"], "=1+0.5")

            binding = get_panel_workbook_binding("26-0036", db_path=self.temp_db)
            self.assertEqual(binding["workbook_path"], os.path.abspath(workbook_path))

    def test_round_trip_save_and_external_conflict_copy(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = os.path.join(temp_dir, "Panels.xlsx")
            self._create_acies_plan_workbook(workbook_path)
            service = PanelScheduleSyncService(
                db_path=self.temp_db,
                prefer_com=False,
            )
            started = service.begin_session("26-0036", workbook_path)

            draft_panels = deepcopy(started["panels"])
            draft_panels[0]["circuits"][0]["loadDescription"] = "APP EDIT"
            saved = service.save_session(started["sessionId"], draft_panels)
            self.assertEqual(saved["status"], "saved")
            self.assertEqual(saved["changedCellCount"], 1)

            check_wb = openpyxl.load_workbook(workbook_path, data_only=False)
            self.assertEqual(check_wb['PNL "LP-1"']["F7"].value, "APP EDIT")
            self.assertEqual(check_wb['PNL "LP-1"']["K7"].value, "=1+0.5")
            check_wb.close()

            clean_external_wb = openpyxl.load_workbook(workbook_path)
            clean_external_wb['PNL "LP-1"']["F7"] = "CLEAN EXTERNAL EDIT"
            clean_external_wb.save(workbook_path)
            clean_external_wb.close()

            auto_refreshed = service.poll_session(saved["sessionId"], is_dirty=False)
            self.assertEqual(auto_refreshed["status"], "updated")
            self.assertEqual(
                auto_refreshed["panels"][0]["circuits"][0]["loadDescription"],
                "CLEAN EXTERNAL EDIT",
            )

            external_wb = openpyxl.load_workbook(workbook_path)
            external_wb['PNL "LP-1"']["F7"] = "EXTERNAL EDIT"
            external_wb.save(workbook_path)
            external_wb.close()

            conflict = service.poll_session(saved["sessionId"], is_dirty=True)
            self.assertEqual(conflict["status"], "conflict")
            self.assertTrue(os.path.exists(conflict["conflictPath"]))
            self.assertTrue(
                any(change["address"] == "F7" for change in conflict["changes"])
            )

            original = openpyxl.load_workbook(workbook_path)
            self.assertEqual(original['PNL "LP-1"']["F7"].value, "EXTERNAL EDIT")
            original.close()
            highlighted = openpyxl.load_workbook(conflict["conflictPath"])
            conflict_cell = highlighted['PNL "LP-1"']["F7"]
            self.assertEqual(conflict_cell.value, "EXTERNAL EDIT")
            self.assertEqual(conflict_cell.fill.fill_type, "solid")
            self.assertIn(conflict_cell.fill.fgColor.rgb, ("00FFF2CC", "FFF2CC"))
            highlighted.close()

            refreshed = service.reload_session(saved["sessionId"])
            self.assertEqual(refreshed["status"], "updated")
            self.assertEqual(
                refreshed["panels"][0]["circuits"][0]["loadDescription"],
                "EXTERNAL EDIT",
            )

    def test_generated_acies_layout_saves_headers_and_both_circuit_sides(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = os.path.join(temp_dir, "Generated Panels.xlsx")
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "R2"
            worksheet["A3"] = "(E) PANEL 'R2'"
            worksheet["G2"] = "277/480V"
            worksheet["G3"] = "225A"
            worksheet["G4"] = "M.L.O."
            worksheet["K2"] = 4
            worksheet["K3"] = 3
            worksheet["K4"] = "NEMA 1"
            worksheet["N2"] = "SURFACE"
            worksheet["N3"] = "22KAIC"
            worksheet["N4"] = "EXISTING"
            worksheet["P3"] = "FED FROM MAIN SERVICE"
            worksheet["A6"] = "CKT#"
            worksheet["S6"] = "CKT#"
            worksheet["A8"] = 1
            worksheet["F8"] = "ODD LOAD"
            worksheet["I8"] = 1.25
            worksheet["J8"] = "A"
            worksheet["K8"] = "=2+3"
            worksheet["L8"] = "EVEN LOAD"
            worksheet["S8"] = 2
            workbook.save(workbook_path)
            workbook.close()

            service = PanelScheduleSyncService(
                db_path=self.temp_db,
                prefer_com=False,
            )
            started = service.begin_session("26-0036", workbook_path)
            panel = started["panels"][0]
            self.assertEqual(panel["layout"], "generated_template_v1")
            self.assertEqual(len(panel["circuits"]), 42)
            self.assertEqual(panel["circuits"][-1]["circuitNumber"], 42)
            self.assertEqual(panel["mainType"], "MLO")
            self.assertEqual(panel["mainBreakerAmps"], 0)
            self.assertEqual(panel["panelType"], "NEMA 1")
            self.assertEqual(panel["condition"], "EXISTING")
            self.assertEqual(panel["feedFrom"], "FED FROM MAIN SERVICE")

            odd = next(circuit for circuit in panel["circuits"] if circuit["circuitNumber"] == 1)
            even = next(circuit for circuit in panel["circuits"] if circuit["circuitNumber"] == 2)
            self.assertEqual(odd["side"], "left")
            self.assertEqual(even["side"], "right")
            self.assertEqual(even["connectedKvaInput"], "=2+3")

            draft_panels = deepcopy(started["panels"])
            draft = draft_panels[0]
            draft["mainType"] = "MCB"
            draft["mainBreakerAmps"] = 175
            draft["panelType"] = "NEMA 3R"
            next(circuit for circuit in draft["circuits"] if circuit["circuitNumber"] == 1)[
                "loadDescription"
            ] = "UPDATED ODD LOAD"
            next(circuit for circuit in draft["circuits"] if circuit["circuitNumber"] == 2)[
                "loadDescription"
            ] = "UPDATED EVEN LOAD"

            saved = service.save_session(started["sessionId"], draft_panels)
            self.assertEqual(saved["status"], "saved")
            self.assertEqual(saved["changedCellCount"], 4)

            check_workbook = openpyxl.load_workbook(workbook_path, data_only=False)
            check_sheet = check_workbook["R2"]
            self.assertEqual(check_sheet["G4"].value, "175A")
            self.assertEqual(check_sheet["K4"].value, "NEMA 3R")
            self.assertEqual(check_sheet["F8"].value, "UPDATED ODD LOAD")
            self.assertEqual(check_sheet["L8"].value, "UPDATED EVEN LOAD")
            self.assertEqual(check_sheet["K8"].value, "=2+3")
            check_workbook.close()


if __name__ == "__main__":
    unittest.main()
