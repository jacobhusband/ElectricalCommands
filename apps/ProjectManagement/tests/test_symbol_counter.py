import io
import tempfile
import unittest
from pathlib import Path

import pymupdf as fitz
from PIL import Image, ImageDraw
from openpyxl import load_workbook

from symbol_counter import SymbolCounterError, SymbolCounterService


def _drawing_page(symbol_centers):
    image = Image.new("L", (500, 400), 255)
    draw = ImageDraw.Draw(image)
    draw.line((20, 120, 480, 120), fill=205, width=2)
    draw.rectangle((40, 40, 460, 360), outline=225, width=2)
    for center_x, center_y in symbol_centers:
        left = center_x - 12
        top = center_y - 12
        right = center_x + 12
        bottom = center_y + 12
        draw.ellipse((left, top, right, bottom), outline=0, width=3)
        draw.line((left + 5, center_y, right - 5, center_y), fill=0, width=3)
        draw.line((center_x, top + 5, center_x, bottom - 5), fill=0, width=3)
        draw.rectangle((center_x - 3, center_y - 3, center_x + 3, center_y + 3), fill=0)
    payload = io.BytesIO()
    image.save(payload, format="PNG")
    return payload.getvalue()


def _create_test_pdf(path):
    document = fitz.open()
    for centers in [[(90, 80), (260, 175)], [(150, 240)]]:
        page = document.new_page(width=500, height=400)
        page.insert_image(page.rect, stream=_drawing_page(centers))
    document.save(path)
    document.close()


class SymbolCounterServiceTests(unittest.TestCase):
    def test_counts_selected_symbol_across_all_pdf_pages(self):
        with tempfile.TemporaryDirectory(prefix="acies-symbol-count-") as temp_dir:
            temp_path = Path(temp_dir)
            pdf_path = temp_path / "Electrical Plans.pdf"
            _create_test_pdf(pdf_path)
            service = SymbolCounterService(temp_path / "cache", render_dpi=120)

            session = service.prepare_documents([pdf_path])
            result = service.count(
                session["sessionId"],
                {
                    "sourcePageId": session["pages"][0]["id"],
                    "selection": {
                        "x": 74 / 500,
                        "y": 64 / 400,
                        "width": 32 / 500,
                        "height": 32 / 400,
                    },
                    "threshold": 0.8,
                    "scope": "all",
                    "rotations": False,
                    "scaleTolerance": False,
                },
            )

            self.assertEqual(3, result["total"])
            self.assertEqual([2, 1], [item["count"] for item in result["pageCounts"]])
            self.assertTrue(all(match["score"] >= 0.8 for match in result["matches"]))
            self.assertGreater(result["templateRect"]["pixelWidth"], 8)

    def test_exports_summary_page_counts_and_audit_trail(self):
        with tempfile.TemporaryDirectory(prefix="acies-symbol-export-") as temp_dir:
            temp_path = Path(temp_dir)
            pdf_path = temp_path / "Power Plan.pdf"
            _create_test_pdf(pdf_path)
            service = SymbolCounterService(temp_path / "cache")
            session = service.prepare_documents([pdf_path])
            page_ids = [page["id"] for page in session["pages"]]
            output_path = temp_path / "takeoff.xlsx"

            exported = service.export_results(
                session["sessionId"],
                [
                    {
                        "name": "Duplex receptacle",
                        "matches": [
                            {
                                "pageId": page_ids[0],
                                "x": 0.1,
                                "y": 0.2,
                                "score": 0.94,
                                "rotation": 0,
                                "scale": 1,
                            },
                            {
                                "pageId": page_ids[1],
                                "x": 0.3,
                                "y": 0.4,
                                "manual": True,
                            },
                        ],
                    }
                ],
                output_path,
            )

            self.assertEqual(2, exported["instanceCount"])
            workbook = load_workbook(output_path, data_only=True)
            self.assertEqual(["Summary", "Counts by Page", "Audit Trail"], workbook.sheetnames)
            self.assertEqual("Duplex receptacle", workbook["Summary"]["A2"].value)
            self.assertEqual(2, workbook["Summary"]["B2"].value)
            self.assertEqual("Manual", workbook["Audit Trail"]["G3"].value)

    def test_rejects_blank_or_tiny_selections(self):
        with tempfile.TemporaryDirectory(prefix="acies-symbol-invalid-") as temp_dir:
            temp_path = Path(temp_dir)
            pdf_path = temp_path / "Plan.pdf"
            _create_test_pdf(pdf_path)
            service = SymbolCounterService(temp_path / "cache")
            session = service.prepare_documents([pdf_path])
            with self.assertRaises(SymbolCounterError):
                service.count(
                    session["sessionId"],
                    {
                        "sourcePageId": session["pages"][0]["id"],
                        "selection": {"x": 0.01, "y": 0.01, "width": 0.001, "height": 0.001},
                    },
                )


class SymbolCounterUiContractTests(unittest.TestCase):
    def test_tool_card_dialog_and_frontend_bundle_are_wired(self):
        project_root = Path(__file__).resolve().parents[1]
        html = (project_root / "index.html").read_text(encoding="utf-8")
        script = (project_root / "symbol_counter_ui.js").read_text(encoding="utf-8")
        css = (project_root / "symbol_counter.css").read_text(encoding="utf-8")

        self.assertIn('id="toolSymbolCounter"', html)
        self.assertIn('id="symbolCounterDlg"', html)
        self.assertIn('id="symbolCounterClearAllBtn"', html)
        self.assertIn('src="symbol_counter_ui.js', html)
        self.assertIn("prepare_symbol_count_documents", script)
        self.assertIn("count_pdf_symbols", script)
        self.assertIn("export_symbol_count_results", script)
        self.assertIn("function clearAllCounts()", script)
        self.assertIn("Delete all ${markerCount} counted marker", script)
        self.assertIn("function removeSymbol(symbolId)", script)
        self.assertIn('className = "symbol-counter-symbol-remove"', script)
        self.assertIn(".symbol-counter-symbol-remove", css)
        self.assertIn(".symbol-counter-dialog", css)


if __name__ == "__main__":
    unittest.main()
