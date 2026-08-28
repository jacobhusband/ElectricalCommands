"""ACIES panel-schedule workbook parsing and bidirectional synchronization.

The production panel schedules are legacy ``.xls`` workbooks linked into
AutoCAD as OLE objects. This module edits workbooks in place through Microsoft
Excel when it is available, preserving worksheet names, formulas, formatting,
and the legacy file format. ``.xlsx`` files have an openpyxl fallback for tests
and non-Excel environments.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import threading
import uuid
from contextlib import contextmanager
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill

try:
    from apps.ProjectManagement.database import (
        get_panel_workbook_binding,
        set_panel_workbook_binding,
        upsert_panel_schedule,
    )
except ImportError:  # pragma: no cover - direct execution fallback
    from database import (  # type: ignore
        get_panel_workbook_binding,
        set_panel_workbook_binding,
        upsert_panel_schedule,
    )


SUPPORTED_EXTENSIONS = {".xls", ".xlsx"}

# Older generated-template constants are retained for compatibility with the
# existing Panel Schedule AI workbook and its tests.
COL_L_NOTE = "B"
COL_L_TYPE = "C"
COL_L_POLE = "D"
COL_L_TRIP = "E"
COL_L_DESC = "F"
COL_L_KVA = "I"
COL_VOLTAGE = "G"
COL_R_KVA = "K"
COL_R_DESC = "L"
COL_R_POLE = "O"
COL_R_TRIP = "P"
COL_R_TYPE = "Q"
COL_R_NOTE = "R"
START_ROW = 8
MAX_ROW = 28

# Attached ACIES electrical-plan workbook layout.
ACIES_START_ROW = 7
ACIES_MAX_ROW = 48
ACIES_LAST_SUMMARY_ROW = 57
ACIES_LAST_COLUMN = 21  # U
ACIES_LEFT_COLUMNS = {
    "circuitNumber": "A",
    "notes": "B",
    "loadTypeCode": "C",
    "polesInput": "D",
    "breakerAmpsInput": "E",
    "loadDescription": "F",
    "connectedKvaInput": "K",
}
ACIES_RIGHT_COLUMNS = {
    "connectedKvaInput": "M",
    "loadDescription": "N",
    "polesInput": "Q",
    "breakerAmpsInput": "R",
    "loadTypeCode": "S",
    "notes": "T",
    "circuitNumber": "U",
}
ACIES_HEADER_CELLS = {
    "panelName": "A2",
    "sectionLabel": "A3",
    "voltagePrimary": "G1",
    "voltageSecondary": "H1",
    "voltageUnit": "I1",
    "mainBusRatingAmps": "G2",
    "mainBusUnit": "H2",
    "mainBreakerAmps": "G3",
    "mainType": "H3",
    "wire": "M1",
    "phase": "M2",
    "panelType": "M3",
    "mounting": "P1",
    "aicRatingText": "P2",
    "condition": "P3",
    "feedLabel": "R2",
    "feedFrom": "R3",
}

# Generated ACIES schedule layout used by the Panel Schedule AI tool and by
# existing production workbooks. Several logical fields intentionally share
# G4 because that cell contains either an MLO label or the main-breaker rating.
GENERATED_HEADER_CELLS = {
    "panelName": "A3",
    "voltage": "G2",
    "mainBusRatingAmps": "G3",
    "mainType": "G4",
    "mainBreakerAmps": "G4",
    "wire": "K2",
    "phase": "K3",
    "panelType": "K4",
    "mounting": "N2",
    "aicRatingText": "N3",
    "condition": "N4",
    "feedFrom": "P3",
}

LEGACY_LEFT_COLUMNS = {
    "circuitNumber": "A",
    "notes": COL_L_NOTE,
    "loadTypeCode": COL_L_TYPE,
    "polesInput": COL_L_POLE,
    "breakerAmpsInput": COL_L_TRIP,
    "loadDescription": COL_L_DESC,
    "connectedKvaInput": COL_L_KVA,
}
LEGACY_RIGHT_COLUMNS = {
    "connectedKvaInput": COL_R_KVA,
    "loadDescription": COL_R_DESC,
    "polesInput": COL_R_POLE,
    "breakerAmpsInput": COL_R_TRIP,
    "loadTypeCode": COL_R_TYPE,
    "notes": COL_R_NOTE,
    "circuitNumber": "S",
}

CIRCUIT_EDIT_FIELDS = (
    "notes",
    "loadTypeCode",
    "polesInput",
    "breakerAmpsInput",
    "loadDescription",
    "connectedKvaInput",
)


class PanelScheduleSyncError(RuntimeError):
    """Raised for workbook or session errors that can be shown to the user."""


def clean_cell_str(val: Any) -> str:
    """Return a compact display string without losing formula text."""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def parse_numeric(val: Any, default: float = 0.0) -> float:
    """Safely extract a numeric value from an Excel value or label."""
    if val is None:
        return default
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    text = str(val).replace(",", "").strip()
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return default
    try:
        return float(match.group(0))
    except ValueError:
        return default


def parse_integer(val: Any, default: int = 0) -> int:
    """Safely extract an integer value."""
    return int(parse_numeric(val, float(default)))


def map_load_type(type_code: str, description: str) -> str:
    """Map schedule shorthand to the shared canonical load-type enum."""
    code = clean_cell_str(type_code).upper()
    desc = clean_cell_str(description).upper()
    if "SPARE" in desc:
        return "SPARE"
    if "SPACE" in desc:
        return "SPACE"
    if code in ("C", "L") or "LIGHT" in desc or "LTG" in desc:
        return "LIGHTING_CONTINUOUS"
    if code in ("R", "G") or any(x in desc for x in ("RECEP", "PLUG", "OUTLET")):
        return "RECEPTACLE_NON_CONTINUOUS"
    if code in ("M", "ML") or any(x in desc for x in ("MOTOR", "PUMP", "FAN")):
        return "MOTOR"
    if code in ("H", "AC", "HVAC") or any(
        x in desc for x in ("A/C", "HVAC", "AHU", "RTU", "CONDENSER")
    ):
        return "HVAC_CONTINUOUS"
    if code == "K" or any(x in desc for x in ("KITCHEN", "HOOD", "OVEN")):
        return "KITCHEN_EQUIPMENT"
    if "HEAT" in desc or "HEATER" in desc:
        return "ELECTRIC_HEATING"
    return "RECEPTACLE_NON_CONTINUOUS"


def calculate_phase_balance(
    circuits: List[Dict[str, Any]], phase_count: int = 3
) -> Dict[str, Any]:
    """Calculate connected load and phase unbalance for the canonical payload."""
    phase_loads = {"A": 0.0, "B": 0.0, "C": 0.0}
    for circuit in circuits:
        phase = clean_cell_str(circuit.get("phasePole") or "A").upper()
        if phase in phase_loads:
            phase_loads[phase] += float(circuit.get("connectedVA", 0.0) or 0.0)

    active_names = ("A", "B", "C") if phase_count == 3 else ("A", "B")
    active_loads = [phase_loads[name] for name in active_names]
    total_connected = sum(active_loads)
    average = total_connected / len(active_loads) if total_connected else 0.0
    unbalance = ((max(active_loads) - min(active_loads)) / average * 100.0) if average else 0.0
    return {
        "phaseAConnectedVA": round(phase_loads["A"], 2),
        "phaseBConnectedVA": round(phase_loads["B"], 2),
        "phaseCConnectedVA": round(phase_loads["C"], 2) if phase_count == 3 else 0.0,
        "totalConnectedVA": round(total_connected, 2),
        "totalConnectedAmps": 0.0,
        "phaseADemandVA": round(phase_loads["A"], 2),
        "phaseBDemandVA": round(phase_loads["B"], 2),
        "phaseCDemandVA": round(phase_loads["C"], 2) if phase_count == 3 else 0.0,
        "totalDemandVA": round(total_connected, 2),
        "totalDemandAmps": 0.0,
        "unbalancePercentage": round(unbalance, 2),
    }


def _column_number(column: str) -> int:
    value = 0
    for char in column.upper():
        value = value * 26 + ord(char) - 64
    return value


def _cell_address(column: str, row: int) -> str:
    return f"{column}{row}"


def _matrix_get(matrix: List[List[Any]], address: str) -> Any:
    match = re.fullmatch(r"([A-Z]+)(\d+)", address.upper())
    if not match:
        return None
    column, row_text = match.groups()
    row_index = int(row_text) - 1
    column_index = _column_number(column) - 1
    if row_index < 0 or row_index >= len(matrix):
        return None
    row = matrix[row_index]
    return row[column_index] if 0 <= column_index < len(row) else None


def _json_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _formula_or_value(
    values: List[List[Any]], formulas: List[List[Any]], address: str
) -> Any:
    formula = _matrix_get(formulas, address)
    if formula is not None:
        return _json_cell_value(formula)
    return _json_cell_value(_matrix_get(values, address))


def _normalize_voltage(*parts: Any) -> str:
    text = " ".join(clean_cell_str(part) for part in parts if clean_cell_str(part))
    numbers = re.findall(r"\d+(?:\.\d+)?", text)
    if len(numbers) >= 2:
        return f"{int(float(numbers[0]))}/{int(float(numbers[1]))}V"
    if len(numbers) == 1:
        return f"{int(float(numbers[0]))}V"
    return text.replace(" ", "")


def _extract_panel_name(title: str, sheet_name: str) -> str:
    text = clean_cell_str(title)
    quoted = re.search(r"\b(?:PANEL|PNL)\s*[\"']([^\"']+)[\"']", text, re.IGNORECASE)
    if quoted:
        return quoted.group(1).strip()
    unquoted = re.search(r"\b(?:PANEL|PNL)\s+(.+)$", text, re.IGNORECASE)
    if unquoted:
        candidate = re.sub(r"^\([^)]*\)\s*", "", unquoted.group(1)).strip(" \"'")
        if candidate:
            return candidate
    quoted_sheet = re.search(r"[\"']([^\"']+)[\"']", clean_cell_str(sheet_name))
    return quoted_sheet.group(1).strip() if quoted_sheet else clean_cell_str(sheet_name)


def _display_panel_name(panel_name: str, section_label: str, sheet_name: str) -> str:
    name = panel_name or sheet_name
    section = clean_cell_str(section_label)
    if section and section.upper() not in name.upper():
        return f"{name} - {section}"
    return name


def _is_attached_acies_layout(values: List[List[Any]]) -> bool:
    return clean_cell_str(_matrix_get(values, "A5")).upper() == "CKT#"


def _build_circuit(
    *,
    values: List[List[Any]],
    formulas: List[List[Any]],
    row: int,
    side: str,
    circuit_number: int,
    phase_pole: str,
    columns: Dict[str, str],
) -> Dict[str, Any]:
    def value_for(field: str) -> Any:
        column = columns.get(field, "")
        return _matrix_get(values, _cell_address(column, row)) if column else None

    def formula_input(field: str) -> Any:
        column = columns.get(field, "")
        if not column:
            return None
        return _formula_or_value(values, formulas, _cell_address(column, row))

    description = clean_cell_str(value_for("loadDescription"))
    type_code = clean_cell_str(value_for("loadTypeCode"))
    kva_value = parse_numeric(value_for("connectedKvaInput"), 0.0)
    connected_va = kva_value * 1000.0 if 0 < abs(kva_value) < 100 else kva_value
    return {
        "circuitNumber": circuit_number,
        "phasePole": phase_pole or "A",
        "loadDescription": description,
        "loadType": map_load_type(type_code, description),
        "loadTypeCode": type_code,
        "notes": clean_cell_str(value_for("notes")),
        "breakerAmps": parse_integer(value_for("breakerAmpsInput"), 0),
        "breakerAmpsInput": clean_cell_str(formula_input("breakerAmpsInput")),
        "poles": parse_integer(value_for("polesInput"), 0),
        "polesInput": clean_cell_str(formula_input("polesInput")),
        "wireGauge": "#12 AWG",
        "conduitSize": '3/4" C',
        "connectedVA": round(connected_va, 2),
        "connectedKva": round(kva_value, 4),
        "connectedKvaInput": clean_cell_str(formula_input("connectedKvaInput")),
        "demandFactorPercent": 100.0,
        "demandVA": round(connected_va, 2),
        "roomOrZone": "",
        "sourceRow": row,
        "side": side,
        "cellMap": {
            field: _cell_address(column, row)
            for field, column in columns.items()
            if column
        },
    }


def _parse_attached_acies_matrix(
    sheet_name: str,
    values: List[List[Any]],
    formulas: List[List[Any]],
    workbook_path: str,
) -> Dict[str, Any]:
    title = clean_cell_str(_matrix_get(values, "A2"))
    section_label = clean_cell_str(_matrix_get(values, "A3"))
    panel_name = _extract_panel_name(title, sheet_name)
    voltage = _normalize_voltage(
        _matrix_get(values, "G1"), _matrix_get(values, "H1"), _matrix_get(values, "I1")
    )
    bus_amps = parse_integer(_matrix_get(values, "G2"), 0)
    main_type = clean_cell_str(_matrix_get(values, "H3")).upper()
    main_breaker_amps = parse_integer(_matrix_get(values, "G3"), 0)
    if main_type not in ("MCB", "MLO"):
        combined_requirement = " ".join(
            filter(None, (clean_cell_str(_matrix_get(values, "G3")), main_type))
        )
        main_type = "MLO" if "MLO" in combined_requirement.upper() else "MCB"
    phase_count = parse_integer(_matrix_get(values, "M2"), 3)
    if phase_count not in (1, 3):
        phase_count = 3
    wire_count = parse_integer(_matrix_get(values, "M1"), 4)
    if wire_count not in (3, 4):
        wire_count = 4

    circuits: List[Dict[str, Any]] = []
    for row in range(ACIES_START_ROW, ACIES_MAX_ROW + 1):
        phase_pole = clean_cell_str(_matrix_get(values, _cell_address("L", row))).upper()
        left_number = parse_integer(_matrix_get(values, _cell_address("A", row)), 0)
        right_number = parse_integer(_matrix_get(values, _cell_address("U", row)), 0)
        if left_number:
            circuits.append(
                _build_circuit(
                    values=values,
                    formulas=formulas,
                    row=row,
                    side="left",
                    circuit_number=left_number,
                    phase_pole=phase_pole,
                    columns=ACIES_LEFT_COLUMNS,
                )
            )
        if right_number:
            circuits.append(
                _build_circuit(
                    values=values,
                    formulas=formulas,
                    row=row,
                    side="right",
                    circuit_number=right_number,
                    phase_pole=phase_pole,
                    columns=ACIES_RIGHT_COLUMNS,
                )
            )
    circuits.sort(key=lambda item: item["circuitNumber"])
    summary = calculate_phase_balance(circuits, phase_count)
    diagnostics: List[str] = []
    validation_status = "VALID"
    if summary["unbalancePercentage"] > 10.0:
        validation_status = "WARNINGS"
        diagnostics.append(
            f"High phase unbalance detected: {summary['unbalancePercentage']}% (Target <= 5.0%)"
        )

    return {
        "worksheetName": sheet_name,
        "scheduleName": sheet_name,
        "displayName": _display_panel_name(panel_name, section_label, sheet_name),
        "panelName": panel_name,
        "headerTitle": title,
        "sectionLabel": section_label,
        "voltage": voltage or "120/208V",
        "phase": phase_count,
        "wire": wire_count,
        "mainBusRatingAmps": bus_amps,
        "mainType": main_type,
        "mainBreakerAmps": main_breaker_amps,
        "panelType": clean_cell_str(_matrix_get(values, "M3")),
        "mounting": clean_cell_str(_matrix_get(values, "P1")),
        "condition": clean_cell_str(_matrix_get(values, "P3")),
        "feedFrom": clean_cell_str(_matrix_get(values, "R3")),
        "aicRatingText": clean_cell_str(_matrix_get(values, "P2")),
        "shortCircuitCurrentRatingAIC": parse_integer(_matrix_get(values, "P2"), 0),
        "enclosureNema": "NEMA 1",
        "location": "",
        "excelWorkbookPath": workbook_path,
        "validationStatus": validation_status,
        "diagnostics": diagnostics,
        "loadSummary": summary,
        "circuits": circuits,
        "layout": "acies_plan_schedule_v1",
        "cellMap": deepcopy(ACIES_HEADER_CELLS),
        "editablePanelFields": [
            "panelName",
            "sectionLabel",
            "voltage",
            "mainBusRatingAmps",
            "mainType",
            "mainBreakerAmps",
            "phase",
            "wire",
            "panelType",
            "mounting",
            "aicRatingText",
            "condition",
            "feedFrom",
        ],
    }


def _parse_generated_legacy_matrix(
    sheet_name: str,
    values: List[List[Any]],
    formulas: List[List[Any]],
    workbook_path: str,
) -> Dict[str, Any]:
    title = clean_cell_str(_matrix_get(values, "A3"))
    panel_name = _extract_panel_name(title, sheet_name)
    raw_voltage = clean_cell_str(_matrix_get(values, "G2"))
    voltage = "277/480V" if ("480" in raw_voltage or "277" in raw_voltage) else (
        "120/240V" if "240" in raw_voltage else "120/208V"
    )
    bus_amps = parse_integer(_matrix_get(values, "G3"), 225)
    main_requirement = clean_cell_str(_matrix_get(values, "G4"))
    normalized_main_requirement = re.sub(r"[^A-Z]", "", main_requirement.upper())
    main_type = "MLO" if "MLO" in normalized_main_requirement else "MCB"
    main_breaker_amps = (
        0 if main_type == "MLO" else parse_integer(main_requirement, bus_amps)
    )
    phase_count = parse_integer(_matrix_get(values, "K3"), 3)
    if phase_count not in (1, 3):
        phase_count = 3
    wire_count = parse_integer(_matrix_get(values, "K2"), 4)
    if wire_count not in (3, 4):
        wire_count = 4
    phases = ("A", "B", "C") if phase_count == 3 else ("A", "B")
    circuits: List[Dict[str, Any]] = []
    for row in range(START_ROW, MAX_ROW + 1):
        phase_pole = phases[(row - START_ROW) % len(phases)]
        circuits.append(
            _build_circuit(
                values=values,
                formulas=formulas,
                row=row,
                side="left",
                circuit_number=(row - START_ROW) * 2 + 1,
                phase_pole=phase_pole,
                columns=LEGACY_LEFT_COLUMNS,
            )
        )
        circuits.append(
            _build_circuit(
                values=values,
                formulas=formulas,
                row=row,
                side="right",
                circuit_number=(row - START_ROW) * 2 + 2,
                phase_pole=phase_pole,
                columns=LEGACY_RIGHT_COLUMNS,
            )
        )
    summary = calculate_phase_balance(circuits, phase_count)
    diagnostics: List[str] = []
    validation_status = "VALID"
    if summary["unbalancePercentage"] > 10.0:
        validation_status = "WARNINGS"
        diagnostics.append(
            f"High phase unbalance detected: {summary['unbalancePercentage']}% (Target <= 5.0%)"
        )
    return {
        "worksheetName": sheet_name,
        "scheduleName": sheet_name,
        "displayName": panel_name,
        "panelName": panel_name,
        "headerTitle": title,
        "sectionLabel": "",
        "voltage": voltage,
        "phase": phase_count,
        "wire": wire_count,
        "mainBusRatingAmps": bus_amps,
        "mainType": main_type,
        "mainBreakerAmps": main_breaker_amps,
        "panelType": clean_cell_str(_matrix_get(values, "K4")),
        "mounting": clean_cell_str(_matrix_get(values, "N2")),
        "condition": clean_cell_str(_matrix_get(values, "N4")),
        "feedFrom": clean_cell_str(_matrix_get(values, "P3")),
        "aicRatingText": clean_cell_str(_matrix_get(values, "N3")),
        "shortCircuitCurrentRatingAIC": parse_integer(_matrix_get(values, "N3"), 10000),
        "enclosureNema": clean_cell_str(_matrix_get(values, "K4")) or "NEMA 1",
        "location": "ELECTRICAL ROOM",
        "excelWorkbookPath": workbook_path,
        "validationStatus": validation_status,
        "diagnostics": diagnostics,
        "loadSummary": summary,
        "circuits": circuits,
        "layout": "generated_template_v1",
        "cellMap": deepcopy(GENERATED_HEADER_CELLS),
        "editablePanelFields": [
            "panelName",
            "voltage",
            "mainBusRatingAmps",
            "mainType",
            "mainBreakerAmps",
            "phase",
            "wire",
            "panelType",
            "mounting",
            "aicRatingText",
            "condition",
            "feedFrom",
        ],
    }


def _parse_sheet_matrix(
    sheet_name: str,
    values: List[List[Any]],
    formulas: List[List[Any]],
    workbook_path: str,
) -> Dict[str, Any]:
    if _is_attached_acies_layout(values):
        return _parse_attached_acies_matrix(sheet_name, values, formulas, workbook_path)
    return _parse_generated_legacy_matrix(sheet_name, values, formulas, workbook_path)


def _worksheet_matrices(ws, data_ws=None) -> Tuple[List[List[Any]], List[List[Any]]]:
    max_rows = max(ACIES_LAST_SUMMARY_ROW, MAX_ROW, int(ws.max_row or 1))
    max_columns = max(ACIES_LAST_COLUMN, int(ws.max_column or 1))
    values: List[List[Any]] = []
    formulas: List[List[Any]] = []
    for row in range(1, max_rows + 1):
        value_row: List[Any] = []
        formula_row: List[Any] = []
        for column in range(1, max_columns + 1):
            formula_cell = ws.cell(row=row, column=column)
            value_cell = data_ws.cell(row=row, column=column) if data_ws is not None else formula_cell
            value_row.append(value_cell.value)
            formula_row.append(formula_cell.value)
        values.append(value_row)
        formulas.append(formula_row)
    return values, formulas


def parse_panel_sheet(ws, workbook_path: str = "") -> Dict[str, Any]:
    """Parse one openpyxl worksheet using either supported ACIES layout."""
    values, formulas = _worksheet_matrices(ws)
    return _parse_sheet_matrix(ws.title, values, formulas, workbook_path)


def _editable_addresses(panel: Dict[str, Any]) -> Iterable[str]:
    if panel.get("layout") == "acies_plan_schedule_v1":
        yield from ACIES_HEADER_CELLS.values()
        for row in range(ACIES_START_ROW, ACIES_MAX_ROW + 1):
            yield _cell_address("L", row)
            for column in ACIES_LEFT_COLUMNS.values():
                if column:
                    yield _cell_address(column, row)
            for column in ACIES_RIGHT_COLUMNS.values():
                if column:
                    yield _cell_address(column, row)
        return
    if panel.get("layout") == "generated_template_v1":
        yield from GENERATED_HEADER_CELLS.values()
    for circuit in panel.get("circuits", []):
        yield from (circuit.get("cellMap") or {}).values()


def _build_sheet_snapshot(
    panel: Dict[str, Any], values: List[List[Any]], formulas: List[List[Any]]
) -> Dict[str, Any]:
    return {
        address: _formula_or_value(values, formulas, address)
        for address in sorted(set(_editable_addresses(panel)))
    }


def _state_revision(sheet_snapshots: Dict[str, Dict[str, Any]]) -> str:
    payload = json.dumps(sheet_snapshots, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _file_stamp(path: str) -> Dict[str, int]:
    stat = os.stat(path)
    return {"size": int(stat.st_size), "mtimeNs": int(stat.st_mtime_ns)}


def _normalize_matrix(value: Any, rows: int, columns: int) -> List[List[Any]]:
    if isinstance(value, tuple):
        result = [list(row) if isinstance(row, tuple) else [row] for row in value]
    else:
        result = [[value]]
    while len(result) < rows:
        result.append([])
    for row in result:
        row.extend([None] * max(columns - len(row), 0))
    return [row[:columns] for row in result[:rows]]


def _release_com_object(value: Any) -> None:
    if value is None:
        return
    try:
        if hasattr(value, "_oleobj_"):
            value._oleobj_.Release()
    except Exception:
        pass


@contextmanager
def _com_initialized():
    pythoncom = None
    try:
        import pythoncom as _pythoncom

        pythoncom = _pythoncom
        pythoncom.CoInitialize()
    except Exception:
        pythoncom = None
    try:
        yield
    finally:
        if pythoncom is not None:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


def _read_workbook_with_excel(path: str) -> Dict[str, Any]:
    try:
        import win32com.client
    except Exception as exc:  # pragma: no cover - depends on Windows runtime
        raise PanelScheduleSyncError(
            "Reading or editing .xls panel schedules requires Microsoft Excel and pywin32."
        ) from exc

    excel = workbook = None
    panels: List[Dict[str, Any]] = []
    snapshots: Dict[str, Dict[str, Any]] = {}
    with _com_initialized():
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            try:
                excel.AutomationSecurity = 3
            except Exception:
                pass
            workbook = excel.Workbooks.Open(
                os.path.abspath(path),
                UpdateLinks=0,
                ReadOnly=True,
                IgnoreReadOnlyRecommended=True,
                AddToMru=False,
            )
            for index in range(1, int(workbook.Worksheets.Count) + 1):
                worksheet = cell_range = None
                try:
                    worksheet = workbook.Worksheets.Item(index)
                    sheet_name = clean_cell_str(worksheet.Name)
                    cell_range = worksheet.Range(
                        worksheet.Cells(1, 1),
                        worksheet.Cells(ACIES_LAST_SUMMARY_ROW, ACIES_LAST_COLUMN),
                    )
                    values = _normalize_matrix(
                        cell_range.Value2, ACIES_LAST_SUMMARY_ROW, ACIES_LAST_COLUMN
                    )
                    formulas = _normalize_matrix(
                        cell_range.Formula, ACIES_LAST_SUMMARY_ROW, ACIES_LAST_COLUMN
                    )
                    if not any(any(value is not None for value in row) for row in values[:6]):
                        continue
                    panel = _parse_sheet_matrix(sheet_name, values, formulas, path)
                    panels.append(panel)
                    snapshots[sheet_name] = _build_sheet_snapshot(panel, values, formulas)
                finally:
                    _release_com_object(cell_range)
                    _release_com_object(worksheet)
        except Exception as exc:
            if isinstance(exc, PanelScheduleSyncError):
                raise
            raise PanelScheduleSyncError(f"Could not read panel schedule workbook: {exc}") from exc
        finally:
            if workbook is not None:
                try:
                    workbook.Close(False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            _release_com_object(workbook)
            _release_com_object(excel)

    return {
        "workbookPath": os.path.abspath(path),
        "panels": panels,
        "snapshot": snapshots,
        "revision": _state_revision(snapshots),
        "fileStamp": _file_stamp(path),
    }


def _read_workbook_with_openpyxl(path: str) -> Dict[str, Any]:
    formula_book = openpyxl.load_workbook(path, data_only=False)
    data_book = openpyxl.load_workbook(path, data_only=True)
    panels: List[Dict[str, Any]] = []
    snapshots: Dict[str, Dict[str, Any]] = {}
    try:
        for sheet_name in formula_book.sheetnames:
            if sheet_name.upper() in ("TEMPLATE", "SUMMARY", "CONFIG"):
                continue
            formula_ws = formula_book[sheet_name]
            data_ws = data_book[sheet_name]
            values, formulas = _worksheet_matrices(formula_ws, data_ws=data_ws)
            if not any(any(value is not None for value in row) for row in values[:6]):
                continue
            panel = _parse_sheet_matrix(sheet_name, values, formulas, path)
            panels.append(panel)
            snapshots[sheet_name] = _build_sheet_snapshot(panel, values, formulas)
    finally:
        formula_book.close()
        data_book.close()
    return {
        "workbookPath": os.path.abspath(path),
        "panels": panels,
        "snapshot": snapshots,
        "revision": _state_revision(snapshots),
        "fileStamp": _file_stamp(path),
    }


def read_panel_workbook_state(path: str, prefer_com: bool = True) -> Dict[str, Any]:
    """Read panels, managed-cell snapshots, and a stable logical revision."""
    workbook_path = os.path.abspath(str(path or "").strip())
    if not workbook_path or not os.path.isfile(workbook_path):
        raise FileNotFoundError(f"Panel workbook not found at: {workbook_path}")
    extension = Path(workbook_path).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise PanelScheduleSyncError("Panel schedule must be an .xls or .xlsx workbook.")
    if extension == ".xls":
        return _read_workbook_with_excel(workbook_path)
    if prefer_com and os.name == "nt":
        try:
            return _read_workbook_with_excel(workbook_path)
        except PanelScheduleSyncError:
            pass
    return _read_workbook_with_openpyxl(workbook_path)


def parse_panel_workbook(workbook_path: str) -> List[Dict[str, Any]]:
    """Parse all supported panel worksheets from an .xls or .xlsx workbook."""
    return read_panel_workbook_state(workbook_path)["panels"]


def _database_panel_payload(panel: Dict[str, Any], workbook_path: str) -> Dict[str, Any]:
    payload = deepcopy(panel)
    payload["panelName"] = panel.get("displayName") or panel.get("panelName") or panel.get("worksheetName")
    payload["excelWorkbookPath"] = workbook_path
    for circuit in payload.get("circuits", []):
        circuit["loadDescription"] = circuit.get("loadDescription") or "SPARE"
        if not circuit.get("poles"):
            circuit["poles"] = 1
        if not circuit.get("breakerAmps"):
            circuit["breakerAmps"] = 20
    return payload


def sync_panel_workbook_to_db(
    project_id: str, workbook_path: str, db_path: Optional[str] = None
) -> List[Dict[str, Any]]:
    """Import every worksheet into the unified database."""
    panels = parse_panel_workbook(workbook_path)
    synced_records = []
    for panel in panels:
        if not panel.get("panelName") or not panel.get("circuits"):
            continue
        synced_records.append(
            upsert_panel_schedule(
                project_id,
                _database_panel_payload(panel, workbook_path),
                db_path=db_path,
            )
        )
    return synced_records


def _semantic_equal(left: Any, right: Any) -> bool:
    if left is None and right in (None, ""):
        return True
    if right is None and left in (None, ""):
        return True
    if isinstance(left, (int, float)) or isinstance(right, (int, float)):
        try:
            return abs(float(left) - float(right)) < 0.000001
        except (TypeError, ValueError):
            pass
    return clean_cell_str(left) == clean_cell_str(right)


def _format_panel_title(existing: str, panel_name: str) -> str:
    name = clean_cell_str(panel_name)
    current = clean_cell_str(existing)
    quoted = re.search(r"(\b(?:PANEL|PNL)\s*[\"'])([^\"']+)([\"'])", current, re.IGNORECASE)
    if quoted:
        return current[: quoted.start(2)] + name + current[quoted.end(2) :]
    marker = re.search(r"\b(?:PANEL|PNL)\b", current, re.IGNORECASE)
    if marker:
        prefix = current[: marker.end()].rstrip()
        return f'{prefix} "{name}"' if name else prefix
    return f'PNL "{name}"' if name else current


def _voltage_writes(voltage: str, cell_map: Dict[str, str]) -> List[Dict[str, Any]]:
    numbers = re.findall(r"\d+(?:\.\d+)?", clean_cell_str(voltage))
    if not numbers:
        return []
    primary = int(float(numbers[0]))
    secondary = int(float(numbers[1])) if len(numbers) > 1 else None
    return [
        {"address": cell_map["voltagePrimary"], "value": f"{primary} /" if secondary else primary},
        {"address": cell_map["voltageSecondary"], "value": secondary},
        {"address": cell_map["voltageUnit"], "value": "V"},
    ]


def _coerce_write_value(field: str, value: Any) -> Any:
    text = clean_cell_str(value)
    if text == "":
        return None
    if field in ("polesInput", "breakerAmpsInput") and re.fullmatch(r"[-+]?\d+", text):
        return int(text)
    if field == "connectedKvaInput":
        if text.startswith("="):
            return text
        if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text):
            return float(text)
    return text


def _panel_write_operations(
    base_panel: Dict[str, Any], next_panel: Dict[str, Any]
) -> List[Dict[str, Any]]:
    layout = clean_cell_str(base_panel.get("layout"))
    if layout not in ("acies_plan_schedule_v1", "generated_template_v1"):
        raise PanelScheduleSyncError(
            f"Worksheet '{base_panel.get('worksheetName')}' is not the supported ACIES plan schedule layout."
        )
    worksheet_name = clean_cell_str(base_panel.get("worksheetName"))
    default_cell_map = (
        ACIES_HEADER_CELLS
        if layout == "acies_plan_schedule_v1"
        else GENERATED_HEADER_CELLS
    )
    cell_map = base_panel.get("cellMap") or default_cell_map
    writes: List[Dict[str, Any]] = []

    def add(address: str, value: Any, formula_allowed: bool = False) -> None:
        writes.append(
            {
                "sheetName": worksheet_name,
                "address": address,
                "value": value,
                "formulaAllowed": formula_allowed,
            }
        )

    if (
        cell_map.get("panelName")
        and not _semantic_equal(base_panel.get("panelName"), next_panel.get("panelName"))
    ):
        add(
            cell_map["panelName"],
            _format_panel_title(base_panel.get("headerTitle", ""), next_panel.get("panelName", "")),
        )
    if (
        cell_map.get("sectionLabel")
        and not _semantic_equal(base_panel.get("sectionLabel"), next_panel.get("sectionLabel"))
    ):
        add(cell_map["sectionLabel"], clean_cell_str(next_panel.get("sectionLabel")) or None)
    if not _semantic_equal(base_panel.get("voltage"), next_panel.get("voltage")):
        if layout == "acies_plan_schedule_v1":
            for item in _voltage_writes(next_panel.get("voltage", ""), cell_map):
                add(item["address"], item["value"])
        else:
            add(cell_map["voltage"], clean_cell_str(next_panel.get("voltage")) or None)
    if not _semantic_equal(
        base_panel.get("mainBusRatingAmps"), next_panel.get("mainBusRatingAmps")
    ):
        bus_amps = parse_integer(next_panel.get("mainBusRatingAmps"), 0)
        if layout == "acies_plan_schedule_v1":
            add(cell_map["mainBusRatingAmps"], bus_amps or None)
            add(cell_map["mainBusUnit"], "A")
        else:
            add(cell_map["mainBusRatingAmps"], f"{bus_amps}A" if bus_amps else None)
    if (
        not _semantic_equal(base_panel.get("mainType"), next_panel.get("mainType"))
        or not _semantic_equal(base_panel.get("mainBreakerAmps"), next_panel.get("mainBreakerAmps"))
    ):
        main_type = clean_cell_str(next_panel.get("mainType")).upper() or "MLO"
        breaker_amps = parse_integer(next_panel.get("mainBreakerAmps"), 0)
        if layout == "acies_plan_schedule_v1":
            add(
                cell_map["mainBreakerAmps"],
                f"{breaker_amps}A" if main_type == "MCB" and breaker_amps else None,
            )
            add(cell_map["mainType"], main_type)
        else:
            add(
                cell_map["mainType"],
                f"{breaker_amps}A" if main_type == "MCB" and breaker_amps else "MLO",
            )
    for field in ("wire", "phase"):
        if cell_map.get(field) and not _semantic_equal(
            base_panel.get(field), next_panel.get(field)
        ):
            add(cell_map[field], parse_integer(next_panel.get(field), 0) or None)
    for field in ("panelType", "mounting", "aicRatingText", "condition", "feedFrom"):
        if cell_map.get(field) and not _semantic_equal(
            base_panel.get(field), next_panel.get(field)
        ):
            add(cell_map[field], clean_cell_str(next_panel.get(field)) or None)

    base_circuits = {
        (int(circuit.get("sourceRow", 0)), clean_cell_str(circuit.get("side"))): circuit
        for circuit in base_panel.get("circuits", [])
    }
    for circuit in next_panel.get("circuits", []):
        key = (int(circuit.get("sourceRow", 0)), clean_cell_str(circuit.get("side")))
        base_circuit = base_circuits.get(key)
        if not base_circuit:
            continue
        circuit_cells = base_circuit.get("cellMap") or {}
        for field in CIRCUIT_EDIT_FIELDS:
            if field not in circuit_cells:
                continue
            if _semantic_equal(base_circuit.get(field), circuit.get(field)):
                continue
            add(
                circuit_cells[field],
                _coerce_write_value(field, circuit.get(field)),
                formula_allowed=(field == "connectedKvaInput"),
            )
    return writes


def _write_with_excel(path: str, writes: List[Dict[str, Any]]) -> None:
    try:
        import win32com.client
    except Exception as exc:  # pragma: no cover - depends on Windows runtime
        raise PanelScheduleSyncError(
            "Editing panel schedules requires Microsoft Excel and pywin32."
        ) from exc
    excel = workbook = None
    with _com_initialized():
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            try:
                excel.AutomationSecurity = 3
            except Exception:
                pass
            workbook = excel.Workbooks.Open(
                os.path.abspath(path),
                UpdateLinks=0,
                ReadOnly=False,
                IgnoreReadOnlyRecommended=True,
                AddToMru=False,
            )
            if bool(workbook.ReadOnly):
                raise PanelScheduleSyncError(
                    "The panel schedule is read-only or locked by another user. Close Excel or wait for the other user to finish, then try again."
                )
            worksheets = {}
            for write in writes:
                sheet_name = write["sheetName"]
                worksheet = worksheets.get(sheet_name)
                if worksheet is None:
                    worksheet = workbook.Worksheets.Item(sheet_name)
                    worksheets[sheet_name] = worksheet
                cell = worksheet.Range(write["address"])
                value = write.get("value")
                if write.get("formulaAllowed") and isinstance(value, str) and value.startswith("="):
                    cell.Formula = value
                else:
                    cell.Value2 = value
                _release_com_object(cell)
            try:
                workbook.Calculate()
            except Exception:
                pass
            workbook.Save()
            for worksheet in worksheets.values():
                _release_com_object(worksheet)
        except PanelScheduleSyncError:
            raise
        except Exception as exc:
            raise PanelScheduleSyncError(f"Could not save panel schedule workbook: {exc}") from exc
        finally:
            if workbook is not None:
                try:
                    workbook.Close(False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            _release_com_object(workbook)
            _release_com_object(excel)


def _write_with_openpyxl(path: str, writes: List[Dict[str, Any]]) -> None:
    workbook = openpyxl.load_workbook(path, data_only=False)
    try:
        for write in writes:
            workbook[write["sheetName"]][write["address"]] = write.get("value")
        try:
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
        except Exception:
            pass
        workbook.save(path)
    finally:
        workbook.close()


def _write_workbook(path: str, writes: List[Dict[str, Any]], prefer_com: bool) -> None:
    if not writes:
        return
    extension = Path(path).suffix.lower()
    if extension == ".xls" or (prefer_com and os.name == "nt"):
        _write_with_excel(path, writes)
        return
    _write_with_openpyxl(path, writes)


def _diff_snapshots(
    base: Dict[str, Dict[str, Any]], current: Dict[str, Dict[str, Any]]
) -> List[Dict[str, Any]]:
    changes: List[Dict[str, Any]] = []
    for sheet_name in sorted(set(base) | set(current)):
        before_cells = base.get(sheet_name, {})
        after_cells = current.get(sheet_name, {})
        for address in sorted(set(before_cells) | set(after_cells)):
            before = before_cells.get(address)
            after = after_cells.get(address)
            if before == after:
                continue
            changes.append(
                {
                    "worksheetName": sheet_name,
                    "address": address,
                    "field": _describe_cell(address),
                    "before": before,
                    "after": after,
                }
            )
    return changes


def _describe_cell(address: str) -> str:
    header_labels = {
        "A2": "Panel name",
        "A3": "Section",
        "G1": "Voltage",
        "H1": "Voltage",
        "I1": "Voltage",
        "G2": "Bus rating",
        "H2": "Bus rating",
        "G3": "Main requirement",
        "H3": "Main requirement",
        "M1": "Wire",
        "M2": "Phase",
        "M3": "Type",
        "P1": "Mounting",
        "P2": "AIC rating",
        "P3": "Condition",
        "R2": "Feed label",
        "R3": "Fed from",
    }
    if address in header_labels:
        return header_labels[address]
    match = re.fullmatch(r"([A-Z]+)(\d+)", address)
    if not match:
        return address
    column, row_text = match.groups()
    circuit_labels = {
        "A": "Odd circuit number",
        "B": "Odd notes",
        "C": "Odd load type",
        "D": "Odd poles",
        "E": "Odd breaker",
        "F": "Odd description",
        "K": "Odd kVA",
        "L": "Phase",
        "M": "Even kVA",
        "N": "Even description",
        "Q": "Even poles",
        "R": "Even breaker",
        "S": "Even load type",
        "T": "Even notes",
        "U": "Even circuit number",
    }
    return f"{circuit_labels.get(column, address)} (row {row_text})"


def _unique_conflict_path(path: str) -> str:
    source = Path(path)
    stamp = datetime.now().strftime("%Y-%m-%d %H%M%S")
    candidate = source.with_name(f"{source.stem} (External Changes {stamp}){source.suffix}")
    counter = 2
    while candidate.exists():
        candidate = source.with_name(
            f"{source.stem} (External Changes {stamp}-{counter}){source.suffix}"
        )
        counter += 1
    return str(candidate)


def _highlight_with_openpyxl(path: str, changes: List[Dict[str, Any]]) -> None:
    workbook = openpyxl.load_workbook(path, data_only=False)
    yellow = PatternFill(fill_type="solid", fgColor="FFF2CC")
    try:
        for change in changes:
            sheet_name = change.get("worksheetName")
            address = change.get("address")
            if sheet_name not in workbook.sheetnames or not address:
                continue
            cell = workbook[sheet_name][address]
            cell.fill = yellow
            cell.font = Font(
                name=cell.font.name,
                size=cell.font.sz,
                bold=cell.font.bold,
                italic=cell.font.italic,
                color="C00000",
            )
        workbook.save(path)
    finally:
        workbook.close()


def _highlight_with_excel(path: str, changes: List[Dict[str, Any]]) -> None:
    try:
        import win32com.client
    except Exception as exc:  # pragma: no cover
        raise PanelScheduleSyncError("Highlighting .xls conflicts requires Microsoft Excel.") from exc
    excel = workbook = None
    with _com_initialized():
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            workbook = excel.Workbooks.Open(
                os.path.abspath(path), UpdateLinks=0, ReadOnly=False, AddToMru=False
            )
            if bool(workbook.ReadOnly):
                raise PanelScheduleSyncError("The conflict copy could not be opened for highlighting.")
            for change in changes:
                worksheet = cell = target = None
                try:
                    worksheet = workbook.Worksheets.Item(change["worksheetName"])
                    cell = worksheet.Range(change["address"])
                    target = cell.MergeArea if bool(cell.MergeCells) else cell
                    target.Interior.Color = 65535  # yellow
                    target.Font.Color = 255  # red
                except Exception:
                    continue
                finally:
                    if target is not cell:
                        _release_com_object(target)
                    _release_com_object(cell)
                    _release_com_object(worksheet)
            workbook.Save()
        finally:
            if workbook is not None:
                try:
                    workbook.Close(False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            _release_com_object(workbook)
            _release_com_object(excel)


def create_highlighted_conflict_copy(
    workbook_path: str,
    changes: List[Dict[str, Any]],
    prefer_com: bool = True,
) -> str:
    """Copy the current external workbook and highlight externally changed cells."""
    conflict_path = _unique_conflict_path(workbook_path)
    shutil.copy2(workbook_path, conflict_path)
    extension = Path(conflict_path).suffix.lower()
    if extension == ".xls" or (prefer_com and os.name == "nt"):
        _highlight_with_excel(conflict_path, changes)
    else:
        _highlight_with_openpyxl(conflict_path, changes)
    return conflict_path


class PanelScheduleSyncService:
    """Own short-lived edit sessions and detect external workbook changes."""

    def __init__(self, *, db_path: Optional[str] = None, prefer_com: bool = True):
        self.db_path = db_path
        self.prefer_com = prefer_com
        self._sessions: Dict[str, Dict[str, Any]] = {}
        self._lock = threading.RLock()

    def get_binding(self, project_id: str) -> Optional[Dict[str, Any]]:
        return get_panel_workbook_binding(clean_cell_str(project_id), db_path=self.db_path)

    def _public_payload(self, session: Dict[str, Any], status: str = "success") -> Dict[str, Any]:
        return {
            "status": status,
            "sessionId": session["sessionId"],
            "projectId": session["projectId"],
            "workbookPath": session["workbookPath"],
            "revision": session["baseRevision"],
            "panels": deepcopy(session["panels"]),
            "lastSyncedAt": session["lastSyncedAt"],
        }

    def _sync_database_best_effort(self, project_id: str, state: Dict[str, Any]) -> None:
        for panel in state.get("panels", []):
            try:
                upsert_panel_schedule(
                    project_id,
                    _database_panel_payload(panel, state["workbookPath"]),
                    db_path=self.db_path,
                )
            except Exception:
                # Some task-list projects predate the unified projects table.
                return

    def begin_session(self, project_id: str, workbook_path: Optional[str] = None) -> Dict[str, Any]:
        project_key = clean_cell_str(project_id)
        if not project_key:
            raise PanelScheduleSyncError("Choose a project before opening Panel Schedules.")
        path = clean_cell_str(workbook_path)
        if not path:
            binding = self.get_binding(project_key)
            path = clean_cell_str((binding or {}).get("workbook_path"))
        if not path:
            return {"status": "unbound", "projectId": project_key}
        state = read_panel_workbook_state(path, prefer_com=self.prefer_com)
        if not state["panels"]:
            raise PanelScheduleSyncError("No panel schedule worksheets were found in the selected workbook.")
        session_id = str(uuid.uuid4())
        now = datetime.now().astimezone().isoformat()
        session = {
            "sessionId": session_id,
            "projectId": project_key,
            "workbookPath": state["workbookPath"],
            "baseRevision": state["revision"],
            "baseSnapshot": state["snapshot"],
            "fileStamp": state["fileStamp"],
            "panels": state["panels"],
            "lastSyncedAt": now,
            "lastConflictRevision": "",
            "lastConflictPath": "",
            "lastConflictChanges": [],
        }
        with self._lock:
            self._sessions[session_id] = session
        set_panel_workbook_binding(project_key, state["workbookPath"], db_path=self.db_path)
        self._sync_database_best_effort(project_key, state)
        return self._public_payload(session)

    def _get_session(self, session_id: str) -> Dict[str, Any]:
        with self._lock:
            session = self._sessions.get(clean_cell_str(session_id))
        if not session:
            raise PanelScheduleSyncError("The panel schedule edit session has expired. Reopen the tool.")
        return session

    def _conflict_payload(
        self, session: Dict[str, Any], current: Dict[str, Any], changes: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        if session.get("lastConflictRevision") != current["revision"]:
            conflict_path = create_highlighted_conflict_copy(
                session["workbookPath"], changes, prefer_com=self.prefer_com
            )
            session["lastConflictRevision"] = current["revision"]
            session["lastConflictPath"] = conflict_path
            session["lastConflictChanges"] = deepcopy(changes)
        return {
            "status": "conflict",
            "sessionId": session["sessionId"],
            "workbookPath": session["workbookPath"],
            "conflictPath": session["lastConflictPath"],
            "changes": deepcopy(session["lastConflictChanges"]),
            "message": (
                "The Excel workbook changed after this edit session began. "
                "The original workbook was not overwritten; a highlighted conflict copy was created."
            ),
        }

    def poll_session(self, session_id: str, *, is_dirty: bool = False) -> Dict[str, Any]:
        session = self._get_session(session_id)
        try:
            current_stamp = _file_stamp(session["workbookPath"])
        except FileNotFoundError as exc:
            raise PanelScheduleSyncError("The linked panel schedule workbook is no longer available.") from exc
        if current_stamp == session["fileStamp"]:
            return {"status": "unchanged", "revision": session["baseRevision"]}
        current = read_panel_workbook_state(
            session["workbookPath"], prefer_com=self.prefer_com
        )
        if current["revision"] == session["baseRevision"]:
            session["fileStamp"] = current["fileStamp"]
            return {"status": "unchanged", "revision": session["baseRevision"]}
        changes = _diff_snapshots(session["baseSnapshot"], current["snapshot"])
        if is_dirty:
            return self._conflict_payload(session, current, changes)
        session.update(
            {
                "baseRevision": current["revision"],
                "baseSnapshot": current["snapshot"],
                "fileStamp": current["fileStamp"],
                "panels": current["panels"],
                "lastSyncedAt": datetime.now().astimezone().isoformat(),
                "lastConflictRevision": "",
                "lastConflictPath": "",
                "lastConflictChanges": [],
            }
        )
        self._sync_database_best_effort(session["projectId"], current)
        return self._public_payload(session, status="updated")

    def reload_session(self, session_id: str) -> Dict[str, Any]:
        session = self._get_session(session_id)
        current = read_panel_workbook_state(
            session["workbookPath"], prefer_com=self.prefer_com
        )
        session.update(
            {
                "baseRevision": current["revision"],
                "baseSnapshot": current["snapshot"],
                "fileStamp": current["fileStamp"],
                "panels": current["panels"],
                "lastSyncedAt": datetime.now().astimezone().isoformat(),
                "lastConflictRevision": "",
                "lastConflictPath": "",
                "lastConflictChanges": [],
            }
        )
        self._sync_database_best_effort(session["projectId"], current)
        return self._public_payload(session, status="updated")

    def save_session(self, session_id: str, panels: List[Dict[str, Any]]) -> Dict[str, Any]:
        session = self._get_session(session_id)
        current = read_panel_workbook_state(
            session["workbookPath"], prefer_com=self.prefer_com
        )
        if current["revision"] != session["baseRevision"]:
            changes = _diff_snapshots(session["baseSnapshot"], current["snapshot"])
            return self._conflict_payload(session, current, changes)

        next_by_sheet = {
            clean_cell_str(panel.get("worksheetName")): panel
            for panel in panels or []
            if clean_cell_str(panel.get("worksheetName"))
        }
        writes: List[Dict[str, Any]] = []
        for base_panel in session["panels"]:
            sheet_name = clean_cell_str(base_panel.get("worksheetName"))
            next_panel = next_by_sheet.get(sheet_name)
            if next_panel is not None:
                writes.extend(_panel_write_operations(base_panel, next_panel))
        _write_workbook(session["workbookPath"], writes, prefer_com=self.prefer_com)
        refreshed = read_panel_workbook_state(
            session["workbookPath"], prefer_com=self.prefer_com
        )
        session.update(
            {
                "baseRevision": refreshed["revision"],
                "baseSnapshot": refreshed["snapshot"],
                "fileStamp": refreshed["fileStamp"],
                "panels": refreshed["panels"],
                "lastSyncedAt": datetime.now().astimezone().isoformat(),
                "lastConflictRevision": "",
                "lastConflictPath": "",
                "lastConflictChanges": [],
            }
        )
        self._sync_database_best_effort(session["projectId"], refreshed)
        payload = self._public_payload(session, status="saved")
        payload["changedCellCount"] = len(writes)
        return payload

    def close_session(self, session_id: str) -> None:
        with self._lock:
            self._sessions.pop(clean_cell_str(session_id), None)
