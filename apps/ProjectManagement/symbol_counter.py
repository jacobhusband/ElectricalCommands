"""PDF symbol counting engine used by the ACIES desktop application.

The service deliberately keeps UI concerns out of the matching layer.  A caller
opens one or more PDFs, renders a page, supplies a normalized selection rectangle,
and receives normalized match rectangles that can be reviewed or edited in the UI.
"""

from __future__ import annotations

import base64
import hashlib
import math
import os
import threading
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

try:
    import pymupdf as fitz
except ImportError:  # pragma: no cover - compatibility with older PyMuPDF builds
    import fitz  # type: ignore

try:
    import cv2
    import numpy as np
except ImportError:  # pragma: no cover - surfaced as an actionable runtime error
    cv2 = None
    np = None


DEFAULT_RENDER_DPI = 120
MIN_SELECTION_PIXELS = 8
MAX_MATCHES_PER_PAGE = 5000


class SymbolCounterError(ValueError):
    """Raised when a symbol-counting request cannot be completed."""


@dataclass(frozen=True)
class _PageRecord:
    page_id: str
    document_index: int
    page_index: int
    document_name: str
    document_path: str
    label: str
    width_points: float
    height_points: float
    pixel_width: int
    pixel_height: int
    cache_path: str

    def public_dict(self) -> dict[str, Any]:
        return {
            "id": self.page_id,
            "documentIndex": self.document_index,
            "pageIndex": self.page_index,
            "documentName": self.document_name,
            "label": self.label,
            "widthPoints": self.width_points,
            "heightPoints": self.height_points,
            "pixelWidth": self.pixel_width,
            "pixelHeight": self.pixel_height,
        }


@dataclass
class _Session:
    session_id: str
    documents: list[dict[str, Any]]
    pages: list[_PageRecord]


def _clean_path(value: Any) -> str:
    return os.path.abspath(os.path.expanduser(str(value or "").strip()))


def _as_float(value: Any, default: float = 0.0) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    return number if math.isfinite(number) else default


def _normalized_rect(value: Any) -> tuple[float, float, float, float]:
    if not isinstance(value, dict):
        raise SymbolCounterError("Draw a box around one complete symbol first.")
    x = _as_float(value.get("x"))
    y = _as_float(value.get("y"))
    width = _as_float(value.get("width"))
    height = _as_float(value.get("height"))
    if width <= 0 or height <= 0:
        raise SymbolCounterError("The symbol selection is empty.")
    x = max(0.0, min(1.0, x))
    y = max(0.0, min(1.0, y))
    width = min(width, 1.0 - x)
    height = min(height, 1.0 - y)
    if width <= 0 or height <= 0:
        raise SymbolCounterError("The symbol selection falls outside the page.")
    return x, y, width, height


def _encode_png(path: str) -> str:
    payload = Path(path).read_bytes()
    return "data:image/png;base64," + base64.b64encode(payload).decode("ascii")


def _read_grayscale(path: str):
    if cv2 is None or np is None:
        raise SymbolCounterError(
            "Symbol matching is unavailable. Install the Project Management requirements and restart the app."
        )
    # np.fromfile + imdecode supports Windows paths containing non-ASCII characters.
    encoded = np.fromfile(path, dtype=np.uint8)
    image = cv2.imdecode(encoded, cv2.IMREAD_GRAYSCALE)
    if image is None:
        raise SymbolCounterError("The rendered drawing page could not be read.")
    return image


def _trim_template(template):
    """Remove excess white border while retaining a small checking margin."""
    ink = 255 - template
    ys, xs = np.where(ink > 18)
    if not len(xs) or not len(ys):
        raise SymbolCounterError("The selected area is blank. Include the symbol linework.")
    padding = 2
    left = max(0, int(xs.min()) - padding)
    top = max(0, int(ys.min()) - padding)
    right = min(template.shape[1], int(xs.max()) + padding + 1)
    bottom = min(template.shape[0], int(ys.max()) + padding + 1)
    trimmed = template[top:bottom, left:right]
    if trimmed.shape[0] < MIN_SELECTION_PIXELS or trimmed.shape[1] < MIN_SELECTION_PIXELS:
        raise SymbolCounterError(
            "The selected symbol is too small. Zoom in and draw a slightly larger box."
        )
    ink_fraction = float(np.count_nonzero(255 - trimmed > 18)) / float(trimmed.size)
    if ink_fraction < 0.01:
        raise SymbolCounterError("The selection contains too little linework to match reliably.")
    return trimmed, (left, top, right, bottom)


def _intersection_over_union(a, b) -> float:
    ax1, ay1, ax2, ay2 = a[:4]
    bx1, by1, bx2, by2 = b[:4]
    ix1 = max(ax1, bx1)
    iy1 = max(ay1, by1)
    ix2 = min(ax2, bx2)
    iy2 = min(ay2, by2)
    intersection = max(0, ix2 - ix1) * max(0, iy2 - iy1)
    if intersection <= 0:
        return 0.0
    union = (ax2 - ax1) * (ay2 - ay1) + (bx2 - bx1) * (by2 - by1) - intersection
    return float(intersection) / float(union) if union > 0 else 0.0


def _non_maximum_suppression(candidates: list[tuple], iou_threshold: float = 0.28):
    """Collapse duplicate responses emitted by rotations and nearby scales."""
    accepted: list[tuple] = []
    for candidate in sorted(candidates, key=lambda item: item[4], reverse=True):
        if any(_intersection_over_union(candidate, prior) >= iou_threshold for prior in accepted):
            continue
        accepted.append(candidate)
        if len(accepted) >= MAX_MATCHES_PER_PAGE:
            break
    return accepted


def _variant_templates(template, allow_rotations: bool, allow_scale_tolerance: bool):
    angles = (0, 90, 180, 270) if allow_rotations else (0,)
    scales = (0.92, 1.0, 1.08) if allow_scale_tolerance else (1.0,)
    variants = []
    seen_shapes = set()
    for angle in angles:
        rotated = template
        if angle == 90:
            rotated = cv2.rotate(template, cv2.ROTATE_90_CLOCKWISE)
        elif angle == 180:
            rotated = cv2.rotate(template, cv2.ROTATE_180)
        elif angle == 270:
            rotated = cv2.rotate(template, cv2.ROTATE_90_COUNTERCLOCKWISE)
        for scale in scales:
            if scale == 1.0:
                variant = rotated
            else:
                width = max(MIN_SELECTION_PIXELS, int(round(rotated.shape[1] * scale)))
                height = max(MIN_SELECTION_PIXELS, int(round(rotated.shape[0] * scale)))
                variant = cv2.resize(rotated, (width, height), interpolation=cv2.INTER_AREA)
            fingerprint = (angle, variant.shape[1], variant.shape[0])
            if fingerprint in seen_shapes:
                continue
            seen_shapes.add(fingerprint)
            variants.append((angle, scale, variant))
    return variants


def _match_page(target_gray, template_gray, threshold: float, rotations: bool, scale_tolerance: bool):
    target_ink = cv2.GaussianBlur(255 - target_gray, (3, 3), 0.45)
    candidates: list[tuple] = []

    for angle, scale, gray_variant in _variant_templates(
        template_gray, rotations, scale_tolerance
    ):
        height, width = gray_variant.shape[:2]
        if height > target_gray.shape[0] or width > target_gray.shape[1]:
            continue
        template_ink = cv2.GaussianBlur(255 - gray_variant, (3, 3), 0.45)
        if float(template_ink.std()) < 1.0:
            continue
        response = cv2.matchTemplate(target_ink, template_ink, cv2.TM_CCOEFF_NORMED)
        peak_width = max(3, min(width, height) // 3)
        peak_height = peak_width
        local_max = cv2.dilate(
            response,
            np.ones((peak_height, peak_width), dtype=np.uint8),
        )
        ys, xs = np.where((response >= threshold) & (response >= local_max - 1e-7))
        if len(xs) > MAX_MATCHES_PER_PAGE:
            scores = response[ys, xs]
            keep = np.argpartition(scores, -MAX_MATCHES_PER_PAGE)[-MAX_MATCHES_PER_PAGE:]
            xs = xs[keep]
            ys = ys[keep]
        for x, y in zip(xs.tolist(), ys.tolist()):
            score = float(response[y, x])
            candidates.append((x, y, x + width, y + height, score, angle, scale))

    return _non_maximum_suppression(candidates)


class SymbolCounterService:
    """Manages short-lived PDF sessions and performs deterministic template matching."""

    def __init__(self, cache_dir: str | os.PathLike[str] | None = None, render_dpi: int = DEFAULT_RENDER_DPI):
        self.cache_dir = Path(cache_dir or (Path.cwd() / ".symbol-counter-cache"))
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self.render_dpi = max(72, min(200, int(render_dpi)))
        self._sessions: dict[str, _Session] = {}
        self._lock = threading.RLock()

    def _require_session(self, session_id: Any) -> _Session:
        key = str(session_id or "").strip()
        with self._lock:
            session = self._sessions.get(key)
        if session is None:
            raise SymbolCounterError("This symbol-counting session has expired. Reload the drawings.")
        return session

    @staticmethod
    def _page_map(session: _Session) -> dict[str, _PageRecord]:
        return {page.page_id: page for page in session.pages}

    def prepare_documents(self, pdf_paths: Iterable[Any]) -> dict[str, Any]:
        unique_paths = []
        seen = set()
        for raw_path in pdf_paths or []:
            path = _clean_path(raw_path)
            if not path or path.lower() in seen:
                continue
            seen.add(path.lower())
            unique_paths.append(path)
        if not unique_paths:
            raise SymbolCounterError("Select at least one PDF drawing set.")

        session_id = uuid.uuid4().hex
        documents: list[dict[str, Any]] = []
        pages: list[_PageRecord] = []
        scale = self.render_dpi / 72.0

        for document_index, pdf_path in enumerate(unique_paths):
            source = Path(pdf_path)
            if source.suffix.lower() != ".pdf":
                raise SymbolCounterError(f"Symbol Counter only accepts PDF files: {source.name}")
            if not source.is_file():
                raise SymbolCounterError(f"Drawing file not found: {source}")
            try:
                document = fitz.open(str(source))
            except Exception as exc:
                raise SymbolCounterError(f"Could not open {source.name}: {exc}") from exc
            try:
                if document.needs_pass:
                    raise SymbolCounterError(f"Password-protected PDF is not supported: {source.name}")
                if document.page_count < 1:
                    raise SymbolCounterError(f"PDF contains no pages: {source.name}")
                stat = source.stat()
                cache_key = hashlib.sha256(
                    f"{source.resolve()}|{stat.st_size}|{stat.st_mtime_ns}|{self.render_dpi}".encode("utf-8")
                ).hexdigest()[:24]
                document_cache = self.cache_dir / cache_key
                document_cache.mkdir(parents=True, exist_ok=True)
                document_pages = []
                for page_index in range(document.page_count):
                    page = document.load_page(page_index)
                    rect = page.rect
                    page_id = f"d{document_index + 1}-p{page_index + 1}"
                    label = str(page.get_label() or "").strip() or f"Page {page_index + 1}"
                    record = _PageRecord(
                        page_id=page_id,
                        document_index=document_index,
                        page_index=page_index,
                        document_name=source.name,
                        document_path=str(source),
                        label=label,
                        width_points=float(rect.width),
                        height_points=float(rect.height),
                        pixel_width=max(1, int(round(rect.width * scale))),
                        pixel_height=max(1, int(round(rect.height * scale))),
                        cache_path=str(document_cache / f"page-{page_index + 1}.png"),
                    )
                    pages.append(record)
                    document_pages.append(page_id)
                documents.append(
                    {
                        "index": document_index,
                        "name": source.name,
                        "path": str(source),
                        "pageCount": document.page_count,
                        "pageIds": document_pages,
                    }
                )
            finally:
                document.close()

        session = _Session(session_id=session_id, documents=documents, pages=pages)
        with self._lock:
            self._sessions[session_id] = session
            # Bound only session metadata; rendered pages are safely reusable on disk.
            while len(self._sessions) > 8:
                oldest_key = next(iter(self._sessions))
                if oldest_key == session_id:
                    break
                self._sessions.pop(oldest_key, None)

        return {
            "sessionId": session_id,
            "documents": documents,
            "pages": [page.public_dict() for page in pages],
            "pageCount": len(pages),
            "renderDpi": self.render_dpi,
        }

    def _render_page(self, page: _PageRecord) -> str:
        target = Path(page.cache_path)
        if target.is_file() and target.stat().st_size > 0:
            return str(target)
        target.parent.mkdir(parents=True, exist_ok=True)
        document = fitz.open(page.document_path)
        temp_path = target.with_name(f".{target.stem}-{uuid.uuid4().hex}.tmp.png")
        try:
            pdf_page = document.load_page(page.page_index)
            pixmap = pdf_page.get_pixmap(
                matrix=fitz.Matrix(self.render_dpi / 72.0, self.render_dpi / 72.0),
                colorspace=fitz.csGRAY,
                alpha=False,
            )
            pixmap.save(str(temp_path))
            os.replace(temp_path, target)
        finally:
            document.close()
            if temp_path.exists():
                temp_path.unlink(missing_ok=True)
        return str(target)

    def get_page(self, session_id: Any, page_id: Any) -> dict[str, Any]:
        session = self._require_session(session_id)
        page = self._page_map(session).get(str(page_id or ""))
        if page is None:
            raise SymbolCounterError("The requested drawing page is not part of this session.")
        image_path = self._render_page(page)
        return {
            "page": page.public_dict(),
            "imageDataUrl": _encode_png(image_path),
        }

    def count(self, session_id: Any, request: dict[str, Any]) -> dict[str, Any]:
        if cv2 is None or np is None:
            raise SymbolCounterError(
                "Symbol matching dependencies are missing. Install the Project Management requirements and restart."
            )
        if not isinstance(request, dict):
            raise SymbolCounterError("The symbol-counting request is invalid.")
        session = self._require_session(session_id)
        page_map = self._page_map(session)
        source_page_id = str(request.get("sourcePageId") or "").strip()
        source_page = page_map.get(source_page_id)
        if source_page is None:
            raise SymbolCounterError("Choose a source drawing page before counting.")

        x, y, width, height = _normalized_rect(request.get("selection"))
        source_image = _read_grayscale(self._render_page(source_page))
        image_height, image_width = source_image.shape[:2]
        left = max(0, min(image_width - 1, int(round(x * image_width))))
        top = max(0, min(image_height - 1, int(round(y * image_height))))
        right = max(left + 1, min(image_width, int(round((x + width) * image_width))))
        bottom = max(top + 1, min(image_height, int(round((y + height) * image_height))))
        if right - left < MIN_SELECTION_PIXELS or bottom - top < MIN_SELECTION_PIXELS:
            raise SymbolCounterError("The symbol selection is too small. Draw a larger box.")

        template, trim = _trim_template(source_image[top:bottom, left:right])
        trim_left, trim_top, trim_right, trim_bottom = trim
        template_rect = {
            "x": (left + trim_left) / image_width,
            "y": (top + trim_top) / image_height,
            "width": (trim_right - trim_left) / image_width,
            "height": (trim_bottom - trim_top) / image_height,
            "pixelWidth": int(trim_right - trim_left),
            "pixelHeight": int(trim_bottom - trim_top),
        }

        threshold = max(0.55, min(0.98, _as_float(request.get("threshold"), 0.82)))
        rotations = request.get("rotations") is True
        scale_tolerance = request.get("scaleTolerance") is True
        scope = str(request.get("scope") or "all").strip().lower()
        requested_page_ids = request.get("pageIds")
        if isinstance(requested_page_ids, list) and requested_page_ids:
            target_pages = [page_map[str(item)] for item in requested_page_ids if str(item) in page_map]
        elif scope == "current":
            target_pages = [source_page]
        else:
            target_pages = session.pages
        if not target_pages:
            raise SymbolCounterError("No drawing pages are selected for matching.")

        matches = []
        page_counts = []
        for target_page in target_pages:
            target_image = _read_grayscale(self._render_page(target_page))
            target_height, target_width = target_image.shape[:2]
            page_matches = _match_page(
                target_image,
                template,
                threshold,
                rotations,
                scale_tolerance,
            )
            for match_index, candidate in enumerate(page_matches):
                match_left, match_top, match_right, match_bottom, score, angle, scale = candidate
                match_key = hashlib.sha1(
                    f"{target_page.page_id}|{match_left}|{match_top}|{match_right}|{match_bottom}".encode("utf-8")
                ).hexdigest()[:16]
                matches.append(
                    {
                        "id": f"match-{match_key}",
                        "pageId": target_page.page_id,
                        "x": match_left / target_width,
                        "y": match_top / target_height,
                        "width": (match_right - match_left) / target_width,
                        "height": (match_bottom - match_top) / target_height,
                        "score": round(float(score), 4),
                        "rotation": int(angle),
                        "scale": round(float(scale), 3),
                        "manual": False,
                        "order": match_index + 1,
                    }
                )
            page_counts.append(
                {
                    "pageId": target_page.page_id,
                    "count": len(page_matches),
                }
            )

        return {
            "sourcePageId": source_page_id,
            "templateRect": template_rect,
            "threshold": threshold,
            "matches": matches,
            "pageCounts": page_counts,
            "total": len(matches),
        }

    def get_export_defaults(self, session_id: Any) -> dict[str, str]:
        session = self._require_session(session_id)
        first = Path(session.documents[0]["path"])
        return {
            "directory": str(first.parent),
            "filename": f"{first.stem} Symbol Count.xlsx",
        }

    def export_results(self, session_id: Any, symbols: Any, output_path: Any) -> dict[str, Any]:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:  # pragma: no cover - dependency is already required by the app
            raise SymbolCounterError("Excel export requires openpyxl.") from exc

        session = self._require_session(session_id)
        page_map = self._page_map(session)
        normalized_symbols = []
        for index, raw_symbol in enumerate(symbols if isinstance(symbols, list) else []):
            if not isinstance(raw_symbol, dict):
                continue
            name = str(raw_symbol.get("name") or "").strip() or f"Symbol {index + 1}"
            matches = [
                item
                for item in (raw_symbol.get("matches") or [])
                if isinstance(item, dict) and str(item.get("pageId") or "") in page_map
            ]
            normalized_symbols.append({"name": name, "matches": matches})
        if not normalized_symbols:
            raise SymbolCounterError("Count at least one symbol before exporting.")

        target = Path(_clean_path(output_path))
        if target.suffix.lower() != ".xlsx":
            target = target.with_suffix(".xlsx")
        target.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        summary.append(["ACIES Symbol Count", "Quantity"])
        for symbol in normalized_symbols:
            summary.append([symbol["name"], len(symbol["matches"])])
        summary.append([])
        summary.append(["Generated", datetime.now().astimezone().isoformat(timespec="seconds")])
        summary.append(["Source PDFs", "; ".join(document["name"] for document in session.documents)])

        by_page = workbook.create_sheet("Counts by Page")
        symbol_names = [symbol["name"] for symbol in normalized_symbols]
        by_page.append(["Drawing", "Page", *symbol_names, "Page Total"])
        for page in session.pages:
            row_counts = []
            for symbol in normalized_symbols:
                row_counts.append(
                    sum(1 for match in symbol["matches"] if str(match.get("pageId")) == page.page_id)
                )
            by_page.append([page.document_name, page.label, *row_counts, sum(row_counts)])

        audit = workbook.create_sheet("Audit Trail")
        audit.append(
            [
                "Symbol",
                "Drawing",
                "Page",
                "X (%)",
                "Y (%)",
                "Confidence",
                "Method",
                "Rotation",
                "Scale",
            ]
        )
        for symbol in normalized_symbols:
            for match in symbol["matches"]:
                page = page_map[str(match.get("pageId"))]
                manual = match.get("manual") is True
                audit.append(
                    [
                        symbol["name"],
                        page.document_name,
                        page.label,
                        round(_as_float(match.get("x")) * 100, 2),
                        round(_as_float(match.get("y")) * 100, 2),
                        "" if manual else round(_as_float(match.get("score")) * 100, 1),
                        "Manual" if manual else "Automatic",
                        int(_as_float(match.get("rotation"))),
                        round(_as_float(match.get("scale"), 1.0), 3),
                    ]
                )

        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for column in worksheet.columns:
                letter = column[0].column_letter
                max_length = max(len(str(cell.value or "")) for cell in column)
                worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 48)
            worksheet.auto_filter.ref = worksheet.dimensions

        workbook.save(target)
        return {
            "path": str(target),
            "symbolCount": len(normalized_symbols),
            "instanceCount": sum(len(symbol["matches"]) for symbol in normalized_symbols),
        }

    def close_session(self, session_id: Any) -> None:
        with self._lock:
            self._sessions.pop(str(session_id or "").strip(), None)


__all__ = ["SymbolCounterError", "SymbolCounterService"]
